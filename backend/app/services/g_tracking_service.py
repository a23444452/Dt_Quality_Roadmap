from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook

EXCEL_PATH = Path(__file__).resolve().parent.parent.parent.parent / "Quality Roadmap G$ database_Fin_BOD_Dt.xlsx"


def get_tracking_data() -> dict:
    wb = load_workbook(EXCEL_PATH, data_only=True)

    # Parse 2026 sheet
    ws = wb["2026"]
    items = []
    for r in range(2, ws.max_row + 1):
        plant = ws.cell(r, 1).value
        if not plant:
            continue
        complete_date = ws.cell(r, 5).value
        planned_date = ws.cell(r, 7).value
        items.append({
            "plant": str(plant).strip(),
            "line": str(ws.cell(r, 2).value or "").strip(),
            "category": str(ws.cell(r, 3).value or "").strip(),
            "status": str(ws.cell(r, 4).value or "").strip(),
            "complete_date": complete_date.strftime("%Y-%m-%d") if isinstance(complete_date, datetime) else None,
            "planned_date": planned_date.strftime("%Y-%m-%d") if isinstance(planned_date, datetime) else None,
            "owner": str(ws.cell(r, 8).value or "").strip(),
            "class": str(ws.cell(r, 9).value or "").strip(),
        })

    # Parse G$ BS monthly (2026)
    ws_monthly = wb["G$ BS monthly (2026)"]
    monthly_targets = []
    for r in range(2, ws_monthly.max_row + 1):
        month = ws_monthly.cell(r, 1).value
        if not month:
            continue
        monthly_targets.append({
            "month": str(month).strip(),
            "num": int(ws_monthly.cell(r, 2).value or 0),
            "budget": round(float(ws_monthly.cell(r, 3).value or 0), 2),
            "stretch": round(float(ws_monthly.cell(r, 4).value or 0), 2),
        })

    # Parse G$ BS plant (2026)
    ws_plant = wb["G$ BS plant (2026)"]
    plant_targets = []
    for r in range(2, ws_plant.max_row + 1):
        plant = ws_plant.cell(r, 1).value
        if not plant:
            continue
        plant_targets.append({
            "plant": str(plant).strip(),
            "budget": int(ws_plant.cell(r, 2).value or 0),
            "stretch": int(ws_plant.cell(r, 3).value or 0),
        })

    # Compute monthly actual cumulative completions
    monthly_actuals: dict[int, int] = {}
    for item in items:
        if item["status"] == "Complete" and item["complete_date"]:
            month_num = int(item["complete_date"].split("-")[1])
            monthly_actuals[month_num] = monthly_actuals.get(month_num, 0) + 1

    cumulative = 0
    for target in monthly_targets:
        cumulative += monthly_actuals.get(target["num"], 0)
        target["actual_cumulative"] = cumulative

    return {
        "items": items,
        "monthly_targets": monthly_targets,
        "plant_targets": plant_targets,
    }

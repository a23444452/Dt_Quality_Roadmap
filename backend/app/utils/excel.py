from io import BytesIO

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation

LIST_HEADERS = ["solution", "defect_type", "station", "plant", "line", "status"]
MATRIX_SOLUTION_COLS = ["solution", "defect_type", "station"]


def _header_row_to_index(ws) -> dict[str, int]:
    """Return a mapping of lowercase header name → 0-based column index."""
    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    return {
        str(cell).strip().lower(): idx
        for idx, cell in enumerate(header_row)
        if cell is not None
    }


def parse_list_format(workbook: Workbook) -> list[dict]:
    """Parse list-format workbook into a list of record dicts."""
    ws = workbook.active
    col_index = _header_row_to_index(ws)

    records = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        # Skip fully empty rows
        if all(cell is None for cell in row):
            continue
        record = {
            "solution": _get_cell(row, col_index, "solution"),
            "defect_type": _get_cell(row, col_index, "defect_type"),
            "station": _get_cell(row, col_index, "station"),
            "plant": _get_cell(row, col_index, "plant"),
            "line": _get_cell(row, col_index, "line"),
            "status": _get_cell(row, col_index, "status"),
        }
        records.append(record)
    return records


def parse_matrix_format(workbook: Workbook) -> list[dict]:
    """Parse matrix-format workbook (solutions as rows, lines as columns) into records."""
    ws = workbook.active
    header_row = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))[0]

    # Find where plant/line columns start (after solution meta columns)
    # Header format: solution | defect_type | station | plant:line | plant:line | ...
    meta_count = 3  # solution, defect_type, station
    line_columns = []
    for idx in range(meta_count, len(header_row)):
        cell_value = header_row[idx]
        if cell_value is None:
            continue
        cell_str = str(cell_value).strip()
        if "|" in cell_str:
            plant_name, line_name = [part.strip() for part in cell_str.split("|", 1)]
            line_columns.append((idx, plant_name, line_name))
        else:
            # Fall back: treat entire header as line name, no plant info
            line_columns.append((idx, None, cell_str))

    records = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if all(cell is None for cell in row):
            continue

        solution_name = row[0] if len(row) > 0 else None
        defect_type = row[1] if len(row) > 1 else None
        station = row[2] if len(row) > 2 else None

        for col_idx, plant_name, line_name in line_columns:
            if col_idx >= len(row):
                continue
            status_value = row[col_idx]
            if status_value is None:
                continue
            records.append({
                "solution": solution_name,
                "defect_type": defect_type,
                "station": station,
                "plant": plant_name,
                "line": line_name,
                "status": str(status_value).strip() if status_value else None,
            })
    return records


def generate_list_export(data: list[dict]) -> Workbook:
    """Generate a list-format workbook from record dicts."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Solution Map"

    _write_header(ws, LIST_HEADERS)

    for record in data:
        ws.append([
            record.get("solution", ""),
            record.get("defect_type", ""),
            record.get("station", ""),
            record.get("plant", ""),
            record.get("line", ""),
            record.get("status", ""),
        ])

    return wb


def generate_matrix_export(data: list[dict], lines: list[dict]) -> Workbook:
    """
    Generate a matrix-format workbook.

    lines: list of dicts with keys: id, name, plant (used as column headers)
    data: same record dicts as generate_list_export
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Solution Map"

    # Build header: solution, defect_type, station, then one col per line
    line_headers = [f"{ln['plant']} | {ln['name']}" for ln in lines]
    headers = MATRIX_SOLUTION_COLS + line_headers
    _write_header(ws, headers)

    # Build lookup: (solution_name, station_name) → {line_name: status}
    # Use plant|line as key for uniqueness
    lookup: dict[tuple, dict[str, str]] = {}
    for record in data:
        key = (record.get("solution"), record.get("station"))
        line_key = f"{record.get('plant')} | {record.get('line')}"
        if key not in lookup:
            lookup[key] = {
                "defect_type": record.get("defect_type", ""),
            }
        lookup[key][line_key] = record.get("status", "")

    for (solution_name, station_name), info in lookup.items():
        row = [solution_name, info.get("defect_type", ""), station_name]
        for ln in lines:
            line_key = f"{ln['plant']} | {ln['name']}"
            row.append(info.get(line_key, ""))
        ws.append(row)

    return wb


def generate_template(
    format: str,
    solutions: list[dict] | None = None,
    tank_lines: list[dict] | None = None,
    statuses: list[dict] | None = None,
) -> bytes:
    """Generate a template workbook with headers, reference data sheets, and dropdown validation.

    Args:
        format: 'list' or 'matrix'
        solutions: list of dicts with keys: name, defect_type, station, process
        tank_lines: list of dicts with keys: name, plant, line_type
        statuses: list of dicts with keys: code, name
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Template"

    # Build unique lists for dropdowns
    unique_defect_types = list(dict.fromkeys([s.get("defect_type", "") for s in (solutions or []) if s.get("defect_type")]))
    unique_stations = list(dict.fromkeys([s.get("station", "") for s in (solutions or []) if s.get("station")]))
    unique_plants = list(dict.fromkeys([tl.get("plant", "") for tl in (tank_lines or []) if tl.get("plant")]))
    unique_line_names = list(dict.fromkeys([tl.get("name", "") for tl in (tank_lines or []) if tl.get("name")]))
    status_codes = [st.get("code", "") for st in (statuses or []) if st.get("code")]

    # Number of data rows to apply validation (100 rows should be enough)
    max_rows = 100

    if format == "matrix":
        # Use actual tank/line names as column headers if provided
        if tank_lines:
            line_headers = [f"{tl['plant']} | {tl['name']}" for tl in tank_lines]
        else:
            line_headers = ["Plant A | Line 1", "Plant A | Line 2"]
        _write_header(ws, MATRIX_SOLUTION_COLS + line_headers)

        # Add example row with instructions
        ws.append(["(Solution name)", "(Defect type)", "(Station)"] + ["" for _ in line_headers])

        # Add data validation for matrix format
        # Column B: Defect Type dropdown
        if unique_defect_types:
            dv_defect = _create_dropdown(unique_defect_types)
            dv_defect.add(f"B3:B{max_rows + 2}")
            ws.add_data_validation(dv_defect)

        # Column C: Station dropdown
        if unique_stations:
            dv_station = _create_dropdown(unique_stations)
            dv_station.add(f"C3:C{max_rows + 2}")
            ws.add_data_validation(dv_station)

        # Status columns (D onwards): Status code dropdown - use single DV for all columns
        if status_codes:
            dv_status = _create_dropdown(status_codes)
            for col_idx in range(4, 4 + len(line_headers)):
                col_letter = _col_idx_to_letter(col_idx)
                dv_status.add(f"{col_letter}3:{col_letter}{max_rows + 2}")
            ws.add_data_validation(dv_status)

    else:
        _write_header(ws, LIST_HEADERS)
        # Add example row with instructions
        ws.append(["(Solution name)", "", "", "", "", ""])

        # Add data validation for list format
        # Column B: Defect Type dropdown
        if unique_defect_types:
            dv_defect = _create_dropdown(unique_defect_types)
            dv_defect.add(f"B3:B{max_rows + 2}")
            ws.add_data_validation(dv_defect)

        # Column C: Station dropdown
        if unique_stations:
            dv_station = _create_dropdown(unique_stations)
            dv_station.add(f"C3:C{max_rows + 2}")
            ws.add_data_validation(dv_station)

        # Column D: Plant dropdown
        if unique_plants:
            dv_plant = _create_dropdown(unique_plants)
            dv_plant.add(f"D3:D{max_rows + 2}")
            ws.add_data_validation(dv_plant)

        # Column E: Tank/Line dropdown (just line names)
        if unique_line_names:
            dv_line = _create_dropdown(unique_line_names)
            dv_line.add(f"E3:E{max_rows + 2}")
            ws.add_data_validation(dv_line)

        # Column F: Status dropdown
        if status_codes:
            dv_status = _create_dropdown(status_codes)
            dv_status.add(f"F3:F{max_rows + 2}")
            ws.add_data_validation(dv_status)

    # Adjust column widths for template sheet
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 15
    if format == "list":
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 25
        ws.column_dimensions['F'].width = 20

    # Create reference sheets if data is provided
    if solutions:
        ws_solutions = wb.create_sheet("Reference - Solutions")
        _write_header(ws_solutions, ["Solution Name", "Defect Type", "Station", "Process"])
        for sol in solutions:
            ws_solutions.append([
                sol.get("name", ""),
                sol.get("defect_type", ""),
                sol.get("station", ""),
                sol.get("process", ""),
            ])
        ws_solutions.column_dimensions['A'].width = 35
        ws_solutions.column_dimensions['B'].width = 20
        ws_solutions.column_dimensions['C'].width = 15
        ws_solutions.column_dimensions['D'].width = 12

    if tank_lines:
        ws_lines = wb.create_sheet("Reference - Tank_Lines")
        _write_header(ws_lines, ["Plant", "Tank/Line Name", "Type"])
        for tl in tank_lines:
            ws_lines.append([
                tl.get("plant", ""),
                tl.get("name", ""),
                tl.get("line_type", ""),
            ])
        ws_lines.column_dimensions['A'].width = 15
        ws_lines.column_dimensions['B'].width = 20
        ws_lines.column_dimensions['C'].width = 10

    if statuses:
        ws_statuses = wb.create_sheet("Reference - Statuses")
        _write_header(ws_statuses, ["Status Code", "Status Name", "Description"])
        status_descriptions = {
            "MP": "Mass Production - 已量產",
            "DEVELOPING": "Developing - 開發中",
            "INITIATION": "Initiation - 啟動中",
            "PLANNED": "Planned - 已規劃",
            "RESOURCE_CONSTRAIN": "Resource Constrain - 資源受限",
            "NO_INTENTION": "No Intention - 無意導入",
            "NA": "Not Applicable - 不適用",
        }
        for st in statuses:
            ws_statuses.append([
                st.get("code", ""),
                st.get("name", ""),
                status_descriptions.get(st.get("code", ""), ""),
            ])
        ws_statuses.column_dimensions['A'].width = 20
        ws_statuses.column_dimensions['B'].width = 20
        ws_statuses.column_dimensions['C'].width = 35

    return _workbook_to_bytes(wb)


def workbook_to_bytes(wb: Workbook) -> bytes:
    return _workbook_to_bytes(wb)


# ─── Internal helpers ─────────────────────────────────────────────────────────

def _create_dropdown(options: list[str]) -> DataValidation:
    """Create a DataValidation object for dropdown list.

    Note: Excel inline list validation has a 255 character limit.
    Options containing commas will have commas replaced with semicolons.
    """
    if not options:
        raise ValueError("Cannot create dropdown with empty options list")

    # Replace commas in option values to prevent breaking the dropdown
    clean_options = [opt.replace(",", ";") for opt in options]
    options_str = ",".join(clean_options)

    if len(options_str) <= 255:
        dv = DataValidation(
            type="list",
            formula1=f'"{options_str}"',
            allow_blank=True,
            showDropDown=False,  # False = show dropdown arrow
        )
    else:
        # Truncate to fit 255 char limit, keeping complete options
        truncated = options_str[:255]
        last_comma = truncated.rfind(",")
        if last_comma > 0:
            truncated = truncated[:last_comma]
        dv = DataValidation(
            type="list",
            formula1=f'"{truncated}"',
            allow_blank=True,
            showDropDown=False,
        )
    dv.error = "Please select a value from the dropdown list"
    dv.errorTitle = "Invalid Input"
    dv.prompt = "Select from list"
    dv.promptTitle = "Available Options"
    return dv


def _col_idx_to_letter(col_idx: int) -> str:
    """Convert 1-based column index to Excel column letter (1=A, 2=B, ..., 27=AA)."""
    result = ""
    while col_idx > 0:
        col_idx, remainder = divmod(col_idx - 1, 26)
        result = chr(65 + remainder) + result
    return result


def _get_cell(row: tuple, col_index: dict[str, int], key: str):
    idx = col_index.get(key)
    if idx is None or idx >= len(row):
        return None
    value = row[idx]
    return str(value).strip() if value is not None else None


def _write_header(ws, headers: list[str]) -> None:
    ws.append(headers)
    bold_font = Font(bold=True)
    fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    for cell in ws[1]:
        cell.font = bold_font
        cell.fill = fill


def _workbook_to_bytes(wb: Workbook) -> bytes:
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()

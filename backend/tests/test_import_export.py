"""Tests for import/export APIs and utilities."""
import io

import pytest
from fastapi.testclient import TestClient
from openpyxl import Workbook, load_workbook
from sqlalchemy import BigInteger, Integer, event, create_engine
from sqlalchemy.orm import sessionmaker

from app.database import Base
from app.dependencies import get_db
from app.main import app
from app.models.defect import DefectCategory, DefectType
from app.models.plant import Plant, TankLine
from app.models.process import Process, Station
from app.models.solution import Solution
from app.models.solution_map import SolutionMap
from app.models.status_definition import StatusDefinition
from app.models.user import User
from app.utils.excel import (
    generate_list_export,
    generate_matrix_export,
    parse_list_format,
    parse_matrix_format,
    workbook_to_bytes,
)
from app.utils.security import create_access_token, hash_password


# ─── Fixtures ─────────────────────────────────────────────────────────────────

def _sqlite_bigint_workaround(target, connection, **kwargs):
    for table in target.tables.values():
        for col in table.columns:
            if col.primary_key and isinstance(col.type, BigInteger):
                col.type = Integer()


@pytest.fixture
def client():
    engine = create_engine(
        "sqlite:///:memory:",
        connect_args={"check_same_thread": False},
    )
    connection = engine.connect()
    event.listen(Base.metadata, "before_create", _sqlite_bigint_workaround)
    Base.metadata.create_all(bind=connection)
    event.remove(Base.metadata, "before_create", _sqlite_bigint_workaround)

    TestSession = sessionmaker(bind=connection)

    def override_get_db():
        session = TestSession()
        try:
            yield session
        finally:
            session.close()

    app.dependency_overrides[get_db] = override_get_db
    yield TestClient(app)
    app.dependency_overrides.clear()
    Base.metadata.drop_all(bind=connection)
    connection.close()


def _create_editor(client) -> tuple[int, dict]:
    db = next(app.dependency_overrides[get_db]())
    user = User(
        username="editor",
        email="editor@example.com",
        password_hash=hash_password("TestPass123"),
        display_name="Editor",
        role="editor",
        status="active",
    )
    db.add(user)
    db.commit()
    db.refresh(user)
    token = create_access_token(user.id, user.role)
    return user.id, {"Authorization": f"Bearer {token}"}


def _setup_scenario(client) -> dict:
    """Set up a minimal scenario: process, station, defect, plant, line, status, solution."""
    db = next(app.dependency_overrides[get_db]())

    process = Process(name="Electrocoat", sort_order=1)
    db.add(process)
    db.flush()

    station = Station(process_id=process.id, name="Dip Tank", sort_order=1)
    db.add(station)
    db.flush()

    category = DefectCategory(name="Surface", sort_order=1)
    db.add(category)
    db.flush()

    defect_type = DefectType(category_id=category.id, name="Scratch", sort_order=1)
    db.add(defect_type)
    db.flush()

    plant = Plant(name="Main Factory", code="MF01", sort_order=1)
    db.add(plant)
    db.flush()

    line = TankLine(plant_id=plant.id, name="Line A", code="LA", sort_order=1)
    db.add(line)
    db.flush()

    status = StatusDefinition(code="OPEN", name="Open", color="#FF0000", sort_order=1)
    db.add(status)
    db.flush()

    solution = Solution(
        defect_type_id=defect_type.id,
        station_id=station.id,
        name="Polish Surface",
        sort_order=1,
    )
    db.add(solution)
    db.commit()
    db.refresh(solution)

    return {
        "process_id": process.id,
        "plant_id": plant.id,
        "line_id": line.id,
        "line_name": line.name,
        "status_id": status.id,
        "status_code": status.code,
        "solution_id": solution.id,
        "solution_name": solution.name,
    }


def _make_list_xlsx(rows: list[list]) -> bytes:
    """Create a list-format xlsx in memory with given data rows."""
    wb = Workbook()
    ws = wb.active
    ws.append(["solution", "defect_type", "station", "plant", "line", "status"])
    for row in rows:
        ws.append(row)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def _make_matrix_xlsx(solution_rows: list[list], line_headers: list[str]) -> bytes:
    """Create a matrix-format xlsx in memory."""
    wb = Workbook()
    ws = wb.active
    ws.append(["solution", "defect_type", "station"] + line_headers)
    for row in solution_rows:
        ws.append(row)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


# ─── Unit tests: parse_list_format ────────────────────────────────────────────

def test_parse_list_format_returns_records():
    wb = Workbook()
    ws = wb.active
    ws.append(["solution", "defect_type", "station", "plant", "line", "status"])
    ws.append(["Polish Surface", "Scratch", "Dip Tank", "Main Factory", "Line A", "OPEN"])
    ws.append(["Sand", "Dent", "Prep", "Main Factory", "Line B", "DONE"])

    records = parse_list_format(wb)

    assert len(records) == 2
    assert records[0]["solution"] == "Polish Surface"
    assert records[0]["status"] == "OPEN"
    assert records[1]["line"] == "Line B"


def test_parse_list_format_skips_empty_rows():
    wb = Workbook()
    ws = wb.active
    ws.append(["solution", "defect_type", "station", "plant", "line", "status"])
    ws.append(["Polish Surface", "Scratch", "Dip Tank", "Main Factory", "Line A", "OPEN"])
    ws.append([None, None, None, None, None, None])  # empty row

    records = parse_list_format(wb)
    assert len(records) == 1


# ─── Unit tests: parse_matrix_format ──────────────────────────────────────────

def test_parse_matrix_format_returns_flattened_records():
    wb = Workbook()
    ws = wb.active
    ws.append(["solution", "defect_type", "station", "Main Factory | Line A", "Main Factory | Line B"])
    ws.append(["Polish Surface", "Scratch", "Dip Tank", "OPEN", "DONE"])

    records = parse_matrix_format(wb)

    assert len(records) == 2
    assert records[0]["solution"] == "Polish Surface"
    assert records[0]["plant"] == "Main Factory"
    assert records[0]["line"] == "Line A"
    assert records[0]["status"] == "OPEN"
    assert records[1]["line"] == "Line B"
    assert records[1]["status"] == "DONE"


def test_parse_matrix_format_skips_none_status_cells():
    wb = Workbook()
    ws = wb.active
    ws.append(["solution", "defect_type", "station", "Main Factory | Line A"])
    ws.append(["Polish Surface", "Scratch", "Dip Tank", None])  # no status

    records = parse_matrix_format(wb)
    assert len(records) == 0


def test_parse_matrix_format_skips_empty_string_status_cells():
    wb = Workbook()
    ws = wb.active
    ws.append(["solution", "defect_type", "station", "Main Factory | Line A", "Main Factory | Line B"])
    ws.append(["Polish Surface", "Scratch", "Dip Tank", "MP", ""])  # Line B is empty string
    ws.append(["Coating Fix", "Bubble", "Spray", "DEVELOPING", "  "])  # Line B is whitespace

    records = parse_matrix_format(wb)
    assert len(records) == 2
    assert records[0]["status"] == "MP"
    assert records[0]["line"] == "Line A"
    assert records[1]["status"] == "DEVELOPING"
    assert records[1]["line"] == "Line A"


# ─── Unit tests: generate_list_export ─────────────────────────────────────────

def test_generate_list_export_produces_valid_xlsx():
    data = [
        {"solution": "Polish Surface", "defect_type": "Scratch", "station": "Dip Tank",
         "plant": "Main Factory", "line": "Line A", "status": "OPEN"},
    ]
    wb = generate_list_export(data)
    assert wb is not None

    ws = wb.active
    # Header row
    header = [cell.value for cell in ws[1]]
    assert "solution" in header
    assert "status" in header

    # Data row
    data_row = [cell.value for cell in ws[2]]
    assert "Polish Surface" in data_row
    assert "OPEN" in data_row


def test_generate_list_export_roundtrip():
    """Export then parse back should yield same data."""
    original = [
        {"solution": "Polish Surface", "defect_type": "Scratch", "station": "Dip Tank",
         "plant": "Main Factory", "line": "Line A", "status": "OPEN"},
        {"solution": "Sand Surface", "defect_type": "Dent", "station": "Prep",
         "plant": "Main Factory", "line": "Line B", "status": "DONE"},
    ]
    wb = generate_list_export(original)
    parsed = parse_list_format(wb)

    assert len(parsed) == 2
    assert parsed[0]["solution"] == "Polish Surface"
    assert parsed[1]["status"] == "DONE"


# ─── Unit tests: generate_matrix_export ───────────────────────────────────────

def test_generate_matrix_export_produces_valid_xlsx():
    data = [
        {"solution": "Polish Surface", "defect_type": "Scratch", "station": "Dip Tank",
         "plant": "Main Factory", "line": "Line A", "status": "OPEN"},
    ]
    lines = [{"id": 1, "name": "Line A", "plant": "Main Factory"}]
    wb = generate_matrix_export(data, lines)

    ws = wb.active
    header = [cell.value for cell in ws[1]]
    assert "solution" in header
    assert "Main Factory | Line A" in header


# ─── API tests: /import ───────────────────────────────────────────────────────

def test_import_preview_returns_counts(client):
    uid, hdrs = _create_editor(client)
    ids = _setup_scenario(client)

    file_bytes = _make_list_xlsx([
        [ids["solution_name"], "Scratch", "Dip Tank", "Main Factory", ids["line_name"], ids["status_code"]],
    ])

    resp = client.post(
        "/api/v1/import-export/import",
        files={"file": ("import.xlsx", file_bytes, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        data={"format": "list"},
        headers=hdrs,
    )
    assert resp.status_code == 200
    data = resp.json()["data"]
    assert data["total_rows"] == 1
    assert data["new_records"] == 1
    assert data["updated_records"] == 0
    assert "import_id" in data


def test_import_preview_reports_invalid_status(client):
    uid, hdrs = _create_editor(client)
    ids = _setup_scenario(client)

    file_bytes = _make_list_xlsx([
        [ids["solution_name"], "Scratch", "Dip Tank", "Main Factory", ids["line_name"], "INVALID_CODE"],
    ])

    resp = client.post(
        "/api/v1/import-export/import",
        files={"file": ("import.xlsx", file_bytes, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        data={"format": "list"},
        headers=hdrs,
    )
    assert resp.status_code == 200
    data = resp.json()["data"]
    assert len(data["errors"]) == 1
    assert data["new_records"] == 0


def test_import_requires_editor_role(client):
    db = next(app.dependency_overrides[get_db]())
    viewer = User(
        username="viewer",
        email="viewer@example.com",
        password_hash=hash_password("TestPass123"),
        display_name="Viewer",
        role="viewer",
        status="active",
    )
    db.add(viewer)
    db.commit()
    db.refresh(viewer)
    token = create_access_token(viewer.id, viewer.role)
    hdrs = {"Authorization": f"Bearer {token}"}

    file_bytes = _make_list_xlsx([])
    resp = client.post(
        "/api/v1/import-export/import",
        files={"file": ("import.xlsx", file_bytes, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        data={"format": "list"},
        headers=hdrs,
    )
    assert resp.status_code == 403


# ─── API tests: /import/confirm ───────────────────────────────────────────────

def test_import_confirm_creates_records(client):
    uid, hdrs = _create_editor(client)
    ids = _setup_scenario(client)

    file_bytes = _make_list_xlsx([
        [ids["solution_name"], "Scratch", "Dip Tank", "Main Factory", ids["line_name"], ids["status_code"]],
    ])

    preview_resp = client.post(
        "/api/v1/import-export/import",
        files={"file": ("import.xlsx", file_bytes, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        data={"format": "list"},
        headers=hdrs,
    )
    import_id = preview_resp.json()["data"]["import_id"]

    confirm_resp = client.post(
        "/api/v1/import-export/import/confirm",
        json={"import_id": import_id},
        headers=hdrs,
    )
    assert confirm_resp.status_code == 200
    result = confirm_resp.json()["data"]
    assert result["created"] == 1
    assert result["updated"] == 0
    assert result["imported"] == 1

    # Verify record was actually created
    db = next(app.dependency_overrides[get_db]())
    sm = db.query(SolutionMap).filter(
        SolutionMap.solution_id == ids["solution_id"],
        SolutionMap.tank_line_id == ids["line_id"],
    ).first()
    assert sm is not None


def test_import_confirm_invalid_id_returns_400(client):
    uid, hdrs = _create_editor(client)

    resp = client.post(
        "/api/v1/import-export/import/confirm",
        json={"import_id": "nonexistent-id"},
        headers=hdrs,
    )
    assert resp.status_code == 400


def test_import_confirm_updates_existing_record(client):
    uid, hdrs = _create_editor(client)
    ids = _setup_scenario(client)

    # Create a second status
    db = next(app.dependency_overrides[get_db]())
    status2 = StatusDefinition(code="DONE", name="Done", color="#00FF00", sort_order=2)
    db.add(status2)
    db.commit()
    db.refresh(status2)

    # First import creates the record
    file_bytes = _make_list_xlsx([
        [ids["solution_name"], "Scratch", "Dip Tank", "Main Factory", ids["line_name"], ids["status_code"]],
    ])
    preview = client.post(
        "/api/v1/import-export/import",
        files={"file": ("import.xlsx", file_bytes, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        data={"format": "list"},
        headers=hdrs,
    )
    client.post(
        "/api/v1/import-export/import/confirm",
        json={"import_id": preview.json()["data"]["import_id"]},
        headers=hdrs,
    )

    # Second import updates the record with new status
    file_bytes2 = _make_list_xlsx([
        [ids["solution_name"], "Scratch", "Dip Tank", "Main Factory", ids["line_name"], "DONE"],
    ])
    preview2 = client.post(
        "/api/v1/import-export/import",
        files={"file": ("import.xlsx", file_bytes2, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        data={"format": "list"},
        headers=hdrs,
    )
    assert preview2.json()["data"]["updated_records"] == 1

    confirm2 = client.post(
        "/api/v1/import-export/import/confirm",
        json={"import_id": preview2.json()["data"]["import_id"]},
        headers=hdrs,
    )
    result2 = confirm2.json()["data"]
    assert result2["updated"] == 1
    assert result2["created"] == 0


# ─── API tests: /export ───────────────────────────────────────────────────────

def test_export_generates_valid_xlsx(client):
    uid, hdrs = _create_editor(client)
    ids = _setup_scenario(client)

    # Create a solution map entry
    db = next(app.dependency_overrides[get_db]())
    sm = SolutionMap(
        solution_id=ids["solution_id"],
        tank_line_id=ids["line_id"],
        status_id=ids["status_id"],
        version=1,
    )
    db.add(sm)
    db.commit()

    resp = client.get("/api/v1/import-export/export?format=list", headers=hdrs)
    assert resp.status_code == 200
    assert resp.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    wb = load_workbook(filename=io.BytesIO(resp.content))
    ws = wb.active
    header = [cell.value for cell in ws[1]]
    assert "solution" in header


def test_export_list_includes_solutions_without_solution_map(client):
    """Regression: solutions with zero solution_map entries must still appear."""
    uid, hdrs = _create_editor(client)
    ids = _setup_scenario(client)
    # Intentionally do NOT create any SolutionMap row.

    resp = client.get("/api/v1/import-export/export?format=list", headers=hdrs)
    assert resp.status_code == 200

    wb = load_workbook(filename=io.BytesIO(resp.content))
    ws = wb.active
    solution_col = [cell.value for cell in ws["A"]]
    assert ids["solution_name"] in solution_col, (
        "Solution with no solution_map entry was dropped from list export"
    )


def test_export_matrix_includes_solutions_without_solution_map(client):
    """Regression: same for matrix format — solution row must appear with empty cells."""
    uid, hdrs = _create_editor(client)
    ids = _setup_scenario(client)

    resp = client.get("/api/v1/import-export/export?format=matrix", headers=hdrs)
    assert resp.status_code == 200

    wb = load_workbook(filename=io.BytesIO(resp.content))
    ws = wb.active
    solution_col = [cell.value for cell in ws["A"]]
    assert ids["solution_name"] in solution_col, (
        "Solution with no solution_map entry was dropped from matrix export"
    )


def test_export_matrix_format(client):
    uid, hdrs = _create_editor(client)
    ids = _setup_scenario(client)

    db = next(app.dependency_overrides[get_db]())
    sm = SolutionMap(
        solution_id=ids["solution_id"],
        tank_line_id=ids["line_id"],
        status_id=ids["status_id"],
        version=1,
    )
    db.add(sm)
    db.commit()

    resp = client.get("/api/v1/import-export/export?format=matrix", headers=hdrs)
    assert resp.status_code == 200

    wb = load_workbook(filename=io.BytesIO(resp.content))
    ws = wb.active
    header = [cell.value for cell in ws[1]]
    assert "solution" in header


# ─── API tests: /template ─────────────────────────────────────────────────────

def test_template_list_format_download(client):
    uid, hdrs = _create_editor(client)

    resp = client.get("/api/v1/import-export/template?format=list", headers=hdrs)
    assert resp.status_code == 200
    assert "attachment" in resp.headers["content-disposition"]

    wb = load_workbook(filename=io.BytesIO(resp.content))
    ws = wb.active
    header = [cell.value for cell in ws[1]]
    assert "solution" in header
    assert "status" in header
    # Template has only header row
    data_rows = list(ws.iter_rows(min_row=2, values_only=True))
    assert all(all(c is None for c in row) for row in data_rows)


def test_template_matrix_format_download(client):
    uid, hdrs = _create_editor(client)

    resp = client.get("/api/v1/import-export/template?format=matrix", headers=hdrs)
    assert resp.status_code == 200

    wb = load_workbook(filename=io.BytesIO(resp.content))
    ws = wb.active
    header = [cell.value for cell in ws[1]]
    assert "solution" in header
    assert any("|" in (h or "") for h in header)
